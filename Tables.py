import os
from os import walk
from openpyxl import Workbook, load_workbook

def getFilesInFolder(path):
    f = []
    names = []
    for (dirpath, dirnames, filenames) in walk(path):
        f.extend(filenames)
        names = filenames
        break
    return names

def checkExcelInFolder(path):
    count = []
    files = getFilesInFolder(path)
    for file in files:
        file = str(file)
        name, extension = os.path.splitext(path+file)
        if extension == ".xlsx":
            count.append(file)
    return count

def getData():
    import requests
    from bs4 import BeautifulSoup
    num = 0
    listItems = 0
    tableList = []
    finalList = []
    url = "https://tennistonic.com/"
    html = requests.get(url).text
    soup = BeautifulSoup(html,'lxml')

    body = soup.find("body")
    divs_pagina = body.find_all("div", class_= "this_week_m_inner")
    for element in divs_pagina:
        divs_inner = element.find_all("div")
        for divs in divs_inner:
            printer = divs.text
            printer = printer.split("\n")
            for item in printer:
                if item.strip():
                    if num >= 7 and item not in tableList:
                        tableList.append(item)
                        listItems += 1
                    if listItems == 6:
                        if tableList not in finalList:
                            finalList.append(tableList)
                        listItems = 0
                        tableList = []
                    else:
                        num += 1
    return finalList

def saveToExcel(name, tableList, number, folder):
    excelsInFolder = checkExcelInFolder(folder)
    if name+".xlsx" not in excelsInFolder:
        print('messi1')
        wb = Workbook()
        ws = wb.active
        ws.title = "Tennis Tables"
        ws.append(["Status","1-Player","Score","2-Player","Place","City"])
        for element in tableList:
            ws.append(element)
        wb.save(name+".xlsx")
    elif name+str(number)+".xlsx" not in excelsInFolder:
        wb = Workbook()
        ws = wb.active
        ws.title = "Tennis Tables"
        ws.append(["Status","1-Player","Score","2-Player","Place","City"])
        for element in tableList:
            ws.append(element)
        wb.save(name+str(number)+".xlsx")
    elif name+str(number)+".xlsx" in excelsInFolder:
        number += 1
        saveToExcel(name, tableList, number, folder)

def main():
    folder = os.path.abspath(os.path.join(os.path.dirname(__file__)))
    tableList = getData()
    saveToExcel("Data", tableList, 1, folder)

if __name__ == '__main__':
    main()
